import requests
import tkinter as tk
from tkinter import Label, Button, filedialog
import os
import openpyxl



def get_steam_price(file, folder):
    print(file)
    print(folder)
    currencys = {"us", "de", "tr", "uk"}
    country ={
    "ar": "Argentine Peso",
    "au": "Australian Dollar",
    "br": "Brazilian Real",
    "uk": "British Pound",
    "ca": "Canadian Dollar",
    "cl": "Chilean Peso",
    "cn": "Chinese Yuan",
    "az": "CIS - U.S. Dollar",
    "co": "Colombian Peso",
    "cr": "Costa Rican Colon",
    "de": "Euro",
    "hk": "Hong Kong Dollar",
    "in": "Indian Rupee",
    "id": "Indonesian Rupiah",
    "il": "Israeli New Shekel",
    "jp": "Japanese Yen",
    "kz": "Kazakhstani Tenge",
    "kw": "Kuwaiti Dinar",
    "my": "Malaysian Ringgit",
    "mx": "Mexican Peso",
    "nz": "New Zealand Dollar",
    "no": "Norwegian Krone",
    "pe": "Peruvian Sol",
    "ph": "Philippine Peso",
    "pl": "Polish Zloty",
    "qa": "Qatari Riyal",
    "ru": "Russian Ruble",
    "sa": "Saudi Riyal",
    "sg": "Singapore Dollar",
    "za": "South African Rand",
    "pk": "South Asia - USD",
    "kr": "South Korean Won",
    "ch": "Swiss Franc",
    "tw": "Taiwan Dollar",
    "th": "Thai Baht",
    "tr": "Turkish Lira",
    "ae": "U.A.E. Dirham",
    "us": "U.S. Dollar",
    "ua": "Ukrainian Hryvnia",
    "uy": "Uruguayan Peso",
    "vn": "Vietnamese Dong"
}
    with open(f"{folder}/steamprices.xlsx", "w") as f:
        f.write("")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    with open(file, "r") as f:
        app_ids = [line.strip() for line in f.readlines()]
    startrow = 1
    for i in app_ids:
        startrow += 1
        sheet.cell(row=startrow, column=1, value=i)
        for num, b in enumerate(currencys, start=3):
            url = f"https://store.steampowered.com/api/appdetails?appids={i}&cc={b}&l={b}"
            response = requests.get(url)
            data = response.json()
            try:
                sheet.cell(row=startrow, column=2, value=data[str(i)]["data"]["name"])
                print( data[str(i)]["data"]["name"], country[f"{b}"], data[str(i)]["data"]["price_overview"]["final_formatted"])
                sheet.cell(row=startrow, column=num, value=data[str(i)]["data"]["price_overview"]["final_formatted"])
                workbook.save('steamprices.xlsx')
            except:
                print("Error id: ", i)
                sheet.cell(row=startrow, column=num, value=f"Error id: {i}")
                break
            
        workbook.save('steamprices.xlsx')

    workbook.save('steamprices.xlsx')
    



class SteamPriceApp:
    def __init__(self, master):
        self.master = master
        self.folder_select = None
        self.file_select = None
        master.title("Steam Price Checker")
        master.geometry("400x200") 
        self.label = Label(master, text="Steam app list:")
        self.label.pack()

        self.file = Button(master, text="Chose .txt", command=self.select_file)
        self.file.pack()

        self.label = Label(master, text="Extract to:")
        self.label.pack()


        self.folder = Button(master, text="Chose Folder", command=self.select_folder)
        self.folder.pack()


        self.folder = Button(master, text="Start", command=self.check_price)
        self.folder.pack()

        self.result_label = Label(master, text="")
        self.result_label.pack()

    def select_file(self):
        file_path = filedialog.askopenfilename(title="Dosya Seç", filetypes=(("Metin Dosyaları", "*.txt"), ("Tüm Dosyalar", "*.*")))
        if file_path:
            self.file_select = file_path
        
    def select_folder(self):
        foldername = filedialog.askdirectory()
        if foldername:
            self.folder_select = foldername

    def check_price(self):
        folder = self.folder_select
        file = self.file_select
        get_steam_price(file, folder)
        self.result_label.config(text=f"Done")

if __name__ == "__main__":
    root = tk.Tk()
    app = SteamPriceApp(root)
    root.mainloop()
